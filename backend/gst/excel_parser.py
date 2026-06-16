"""
GST Excel / CSV Bulk Upload Parser
Reads invoices from standard Excel templates and converts to Invoice objects.

Supported formats:
1. TaxCompliance Platform template (our own Excel format)
2. Tally export format
3. Busy export format
4. Generic format (auto-detect columns)

Also handles:
- PDF invoice OCR (via pytesseract / pdfplumber)
- Auto-detect HSN code from description (AI-powered)
- Tax computation from invoice value if taxes not provided
"""
import io
import re
import logging
from typing import Optional
from dataclasses import dataclass

import pandas as pd
import openpyxl
from decimal import Decimal, ROUND_HALF_UP

from .invoice_template import Invoice, InvoiceLine, TaxComputer
from .hsn_master import get_gst_rate, suggest_hsn

logger = logging.getLogger(__name__)


# ─── Column name aliases (maps different Excel headers to standard names) ─────

COLUMN_ALIASES = {
    "invoice_number": [
        "Invoice No", "Invoice Number", "Inv No", "Bill No", "Bill Number",
        "Voucher No", "Voucher Number", "Doc No", "Document No",
    ],
    "invoice_date": [
        "Invoice Date", "Date", "Bill Date", "Voucher Date", "Doc Date",
    ],
    "receiver_gstin": [
        "Party GSTIN", "Customer GSTIN", "Buyer GSTIN", "GSTIN", "GSTN",
        "Receiver GSTIN", "Client GSTIN",
    ],
    "receiver_name": [
        "Party Name", "Customer Name", "Buyer Name", "Client Name",
        "Ledger Name", "Account Name",
    ],
    "place_of_supply": [
        "Place of Supply", "POS", "State", "State Code", "Supply State",
    ],
    "invoice_value": [
        "Invoice Value", "Total Amount", "Gross Amount", "Bill Amount",
        "Total", "Amount", "Invoice Total",
    ],
    "taxable_value": [
        "Taxable Value", "Taxable Amount", "Assessable Value", "Basic Amount",
        "Taxable", "Basic Value",
    ],
    "gst_rate": [
        "GST Rate", "Tax Rate", "Rate", "GST %", "Tax %", "Rate %",
    ],
    "igst": [
        "IGST", "IGST Amount", "Integrated Tax",
    ],
    "cgst": [
        "CGST", "CGST Amount", "Central Tax",
    ],
    "sgst": [
        "SGST", "SGST Amount", "UTGST", "State Tax",
    ],
    "cess": [
        "CESS", "Cess Amount", "Compensation Cess",
    ],
    "hsn_sac": [
        "HSN/SAC", "HSN Code", "SAC Code", "HSN", "SAC", "Commodity Code",
    ],
    "description": [
        "Description", "Item Name", "Product", "Goods/Services",
        "Particulars", "Narration",
    ],
    "quantity": [
        "Quantity", "Qty", "Units", "Nos",
    ],
    "unit": [
        "Unit", "UOM", "UQC", "Unit of Measure",
    ],
    "invoice_type": [
        "Invoice Type", "Supply Type", "Type",
    ],
    "reverse_charge": [
        "Reverse Charge", "RCM", "Reverse Charge Applicable",
    ],
}


class ExcelParser:
    """
    Parses Excel/CSV files into Invoice objects.
    Auto-detects format and normalizes column names.
    """

    def __init__(self, supplier_gstin: str, supplier_state: str):
        self.supplier_gstin = supplier_gstin
        self.supplier_state = supplier_state

    def parse_file(self, file_bytes: bytes, filename: str) -> dict:
        """
        Parse uploaded file. Returns invoices + errors.
        """
        ext = filename.lower().rsplit(".", 1)[-1]
        try:
            if ext in ("xlsx", "xls"):
                df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
            elif ext == "csv":
                df = pd.read_csv(io.BytesIO(file_bytes), dtype=str)
            else:
                return {"error": f"Unsupported format: {ext}. Use .xlsx or .csv"}
        except Exception as e:
            return {"error": f"Failed to read file: {e}"}

        df = df.dropna(how="all")   # Remove completely empty rows
        df = df.fillna("")

        # Normalize column names
        col_map = self._map_columns(df.columns.tolist())
        if not col_map.get("invoice_number"):
            return {"error": "Could not find Invoice Number column. Check template format."}

        df = df.rename(columns=col_map)

        invoices = []
        errors = []

        for idx, row in df.iterrows():
            try:
                inv = self._row_to_invoice(row, idx + 2)   # +2 for header row and 1-indexed
                if inv:
                    invoices.append(inv)
            except Exception as e:
                errors.append({"row": idx + 2, "error": str(e)})

        return {
            "invoices": invoices,
            "parse_errors": errors,
            "total_rows": len(df),
            "parsed_invoices": len(invoices),
            "error_rows": len(errors),
        }

    def _map_columns(self, columns: list[str]) -> dict:
        """Map raw column names to standard field names."""
        col_map = {}
        normalized_cols = {c.strip(): c for c in columns}

        for std_name, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                # Case-insensitive match
                for raw_col in normalized_cols:
                    if raw_col.lower() == alias.lower():
                        col_map[normalized_cols[raw_col]] = std_name
                        break

        return col_map

    def _row_to_invoice(self, row: pd.Series, row_num: int) -> Optional[Invoice]:
        """Convert a DataFrame row to an Invoice object."""
        inv_num = str(row.get("invoice_number", "")).strip()
        if not inv_num:
            return None   # Skip empty rows

        inv_date = self._parse_date(str(row.get("invoice_date", "")))
        receiver_gstin = str(row.get("receiver_gstin", "")).strip().upper()
        receiver_name = str(row.get("receiver_name", "")).strip()
        pos = self._parse_pos(str(row.get("place_of_supply", "")), receiver_gstin)
        reverse_charge = "Y" if str(row.get("reverse_charge", "N")).upper() in ("Y", "YES", "TRUE") else "N"

        # Determine invoice type
        inv_type = self._determine_inv_type(receiver_gstin, pos, row)

        # Parse amounts
        invoice_value = self._to_float(row.get("invoice_value", 0))
        taxable_value = self._to_float(row.get("taxable_value", 0))
        gst_rate = self._to_float(row.get("gst_rate", 0))
        igst = self._to_float(row.get("igst", 0))
        cgst = self._to_float(row.get("cgst", 0))
        sgst = self._to_float(row.get("sgst", 0))
        cess = self._to_float(row.get("cess", 0))

        # Auto-compute taxes if not provided
        if taxable_value > 0 and gst_rate > 0 and igst == 0 and cgst == 0:
            computed = TaxComputer.compute(
                taxable_value, gst_rate, self.supplier_state, pos
            )
            igst = computed["igst"]
            cgst = computed["cgst"]
            sgst = computed["sgst"]

        # Auto-compute taxable value from invoice value if not provided
        if invoice_value > 0 and taxable_value == 0 and gst_rate > 0:
            computed = TaxComputer.compute_from_invoice_value(
                invoice_value, gst_rate, self.supplier_state, pos
            )
            taxable_value = computed["taxable_value"]
            igst = computed["igst"]
            cgst = computed["cgst"]
            sgst = computed["sgst"]

        # Invoice value = taxable + all taxes
        if invoice_value == 0:
            invoice_value = taxable_value + igst + cgst + sgst + cess

        # Build line item
        hsn = str(row.get("hsn_sac", "")).strip()
        description = str(row.get("description", "")).strip()
        quantity = self._to_float(row.get("quantity", 1))
        unit = str(row.get("unit", "NOS")).strip() or "NOS"

        line = InvoiceLine(
            hsn_sac=hsn,
            description=description,
            quantity=quantity,
            unit=unit,
            unit_price=round(taxable_value / quantity, 2) if quantity else 0,
            taxable_value=taxable_value,
            gst_rate=gst_rate,
            igst_amount=igst,
            cgst_amount=cgst,
            sgst_amount=sgst,
            cess_amount=cess,
        )

        return Invoice(
            invoice_number=inv_num,
            invoice_date=inv_date,
            invoice_type=inv_type,
            receiver_gstin=receiver_gstin or None,
            receiver_name=receiver_name,
            place_of_supply=pos,
            invoice_value=invoice_value,
            taxable_value=taxable_value,
            igst_amount=igst,
            cgst_amount=cgst,
            sgst_amount=sgst,
            cess_amount=cess,
            reverse_charge=reverse_charge,
            items=[line],
        )

    def _parse_date(self, date_str: str) -> str:
        """Normalize various date formats to DD/MM/YYYY."""
        date_str = date_str.strip()
        for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%Y"]:
            try:
                from datetime import datetime
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime("%d/%m/%Y")
            except ValueError:
                continue
        return date_str  # Return as-is if no format matches

    def _parse_pos(self, pos_str: str, gstin: str) -> str:
        """Extract 2-digit state code from POS or GSTIN."""
        # Extract from GSTIN first (most reliable)
        if gstin and len(gstin) >= 2 and gstin[:2].isdigit():
            return gstin[:2]

        pos_str = pos_str.strip()
        if pos_str.isdigit() and len(pos_str) == 2:
            return pos_str

        # Map state names to codes
        from .hsn_master import GST_STATE_CODES
        state_name_map = {v.lower(): k for k, v in GST_STATE_CODES.items()}
        if pos_str.lower() in state_name_map:
            return state_name_map[pos_str.lower()]

        return self.supplier_state   # Default to supplier state

    def _determine_inv_type(self, gstin: str, pos: str, row: pd.Series) -> str:
        inv_type_raw = str(row.get("invoice_type", "")).strip().upper()
        if inv_type_raw in ("B2B", "B2CS", "B2CL", "EXP", "CDNR", "CDNUR"):
            return inv_type_raw

        if gstin:
            return "B2B"
        # Unregistered: check if inter-state and > 2.5L
        inv_value = self._to_float(row.get("invoice_value", 0))
        if pos != self.supplier_state and inv_value > 250000:
            return "B2CL"
        return "B2CS"

    @staticmethod
    def _to_float(val) -> float:
        try:
            return float(str(val).replace(",", "").strip() or 0)
        except (ValueError, TypeError):
            return 0.0


class InvoiceOCRParser:
    """
    Extract invoice data from PDF files using pdfplumber + regex.
    Falls back to pytesseract for scanned PDFs.
    """

    GSTIN_PATTERN = re.compile(r'\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}\b')
    AMOUNT_PATTERN = re.compile(r'(?:Rs\.?|₹|INR)\s*([0-9,]+(?:\.[0-9]{1,2})?)')
    DATE_PATTERN = re.compile(r'\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b')
    INV_NO_PATTERN = re.compile(r'(?:Invoice\s*No\.?|Inv\.?\s*No\.?|Bill\s*No\.?)\s*[:#]?\s*([A-Z0-9/\-]+)', re.I)

    def parse_pdf(self, pdf_bytes: bytes) -> dict:
        """Extract invoice fields from PDF using pdfplumber."""
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                text = "\n".join(page.extract_text() or "" for page in pdf.pages)

            return self._extract_from_text(text)

        except Exception as e:
            logger.warning(f"pdfplumber failed: {e}, trying OCR")
            return self._ocr_fallback(pdf_bytes)

    def _extract_from_text(self, text: str) -> dict:
        result = {}

        # Invoice number
        inv_match = self.INV_NO_PATTERN.search(text)
        if inv_match:
            result["invoice_number"] = inv_match.group(1).strip()

        # Dates
        dates = self.DATE_PATTERN.findall(text)
        if dates:
            result["invoice_date"] = dates[0]

        # GSTINs
        gstins = self.GSTIN_PATTERN.findall(text)
        if len(gstins) >= 1:
            result["supplier_gstin"] = gstins[0]
        if len(gstins) >= 2:
            result["receiver_gstin"] = gstins[1]

        # Amounts — find largest amount as invoice total
        amounts = [
            float(a.replace(",", ""))
            for a in self.AMOUNT_PATTERN.findall(text)
        ]
        if amounts:
            result["invoice_value"] = max(amounts)

        # Tax amounts
        for tax_name, pattern in [
            ("igst", r'IGST\s*(?:@\s*[\d.]+%\s*)?\s*[:\-]?\s*([\d,]+\.?\d*)'),
            ("cgst", r'CGST\s*(?:@\s*[\d.]+%\s*)?\s*[:\-]?\s*([\d,]+\.?\d*)'),
            ("sgst", r'(?:SGST|UTGST)\s*(?:@\s*[\d.]+%\s*)?\s*[:\-]?\s*([\d,]+\.?\d*)'),
        ]:
            match = re.search(pattern, text, re.I)
            if match:
                result[tax_name] = float(match.group(1).replace(",", ""))

        # HSN
        hsn_match = re.search(r'(?:HSN|SAC)\s*(?:Code)?\s*[:\-]?\s*(\d{4,8})', text, re.I)
        if hsn_match:
            result["hsn_sac"] = hsn_match.group(1)

        result["parse_method"] = "pdfplumber_text"
        return result

    def _ocr_fallback(self, pdf_bytes: bytes) -> dict:
        """Use Pillow + pytesseract for scanned PDF."""
        try:
            import pytesseract
            from PIL import Image
            import fitz   # PyMuPDF

            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            text = ""
            for page in doc:
                pix = page.get_pixmap(dpi=300)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                text += pytesseract.image_to_string(img, lang="eng")

            result = self._extract_from_text(text)
            result["parse_method"] = "ocr_pytesseract"
            return result
        except Exception as e:
            return {"error": f"OCR failed: {e}"}


def generate_excel_template(supplier_gstin: str) -> bytes:
    """
    Generate the standard TaxCompliance Excel template for invoice upload.
    Returns bytes of .xlsx file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "Invoice No *", "Invoice Date * (DD/MM/YYYY)", "Invoice Type",
        "Receiver GSTIN", "Receiver Name *", "Place of Supply (State Code)",
        "Invoice Value *", "Taxable Value *", "GST Rate % *",
        "IGST", "CGST", "SGST", "CESS",
        "HSN/SAC Code", "Description", "Quantity", "Unit",
        "Reverse Charge (Y/N)",
    ]

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    # Sample rows
    samples = [
        ["INV-001", "01/04/2024", "B2B", "27AABCS1429B1ZB", "ABC Pvt Ltd", "27",
         "11800", "10000", "18", "1800", "", "", "", "8471", "Computer", "1", "NOS", "N"],
        ["INV-002", "05/04/2024", "B2CS", "", "Walk-in Customer", "27",
         "5900", "5000", "18", "", "450", "450", "", "9983", "Consulting Services", "1", "NOS", "N"],
        ["INV-003", "10/04/2024", "EXP", "", "Foreign Customer", "96",
         "50000", "50000", "0", "0", "", "", "", "8517", "Mobile Phones", "10", "NOS", "N"],
    ]

    note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for row_idx, sample in enumerate(samples, 2):
        for col_idx, value in enumerate(sample, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = note_fill
            cell.border = border

    # Add instructions sheet
    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        ("FIELD", "DESCRIPTION", "MANDATORY?", "VALID VALUES"),
        ("Invoice No", "Unique invoice number (max 16 chars)", "YES", "Any alphanumeric"),
        ("Invoice Date", "Date of invoice", "YES", "DD/MM/YYYY format"),
        ("Invoice Type", "Supply type", "NO", "B2B, B2CS, B2CL, EXP, CDNR, CDNUR"),
        ("Receiver GSTIN", "Buyer's GSTIN (15 chars)", "For B2B", "e.g. 27AABCS1429B1ZB"),
        ("Place of Supply", "State code", "YES", "2-digit code e.g. 27=Maharashtra"),
        ("Invoice Value", "Total invoice amount including tax", "YES", "Numeric"),
        ("Taxable Value", "Pre-tax value", "YES", "Numeric"),
        ("GST Rate %", "GST rate applicable", "YES", "0, 5, 12, 18, 28"),
        ("IGST/CGST/SGST", "Tax amounts — leave blank to auto-compute", "NO", "Numeric"),
        ("HSN/SAC Code", "HSN for goods, SAC for services", "YES (> 5Cr turnover)", "4-8 digit code"),
        ("Reverse Charge", "Is RCM applicable?", "NO", "Y or N"),
    ]
    for row in instructions:
        ws_inst.append(row)

    # HSN master sheet
    ws_hsn = wb.create_sheet("HSN Master")
    ws_hsn.append(["HSN Code", "Description", "GST Rate %"])
    from .hsn_master import COMMON_HSN
    for code, data in list(COMMON_HSN.items())[:50]:
        ws_hsn.append([code, data["desc"], data["igst"]])

    # State codes sheet
    ws_state = wb.create_sheet("State Codes")
    ws_state.append(["State Code", "State Name"])
    from .hsn_master import GST_STATE_CODES
    for code, name in GST_STATE_CODES.items():
        ws_state.append([code, name])

    # Set column widths
    col_widths = [15, 22, 15, 20, 25, 20, 15, 15, 12, 12, 12, 12, 12, 15, 25, 10, 8, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
